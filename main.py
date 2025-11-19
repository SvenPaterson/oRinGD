import os
import sys
import datetime
import tempfile
import json
import argparse
from dataclasses import dataclass
from typing import List, Optional, Tuple, Literal, cast

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QPushButton,
    QHeaderView,
    QFileDialog,
    QTextEdit,
    QDialog,
    QLabel,
    QAbstractItemView,
    QTabWidget,
    QSizePolicy,
    QSplitter,
)

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from rating import compute_metrics, table_values, assign_iso23936_rating, Crack

from canvas_gv import CanvasScene, CanvasView


DEFAULT_LAYOUT = {
    "window": {"size": [1000, 1070]},
    "top_splitter": [745, 230],
    "main_splitter": [665, 350],
}


@dataclass
class SessionAnalysis:
    index: int
    image_name: str
    image_path: str
    completed_at: datetime.datetime
    crack_count: int
    total_pct: float
    rating: int
    result: str
    cracks: List[Crack]
    snapshot_png: Optional[bytes] = None


RATING_METRICS = [
    "Total crack length (% of CSD)",
    "# cracks that are <25% CSD",
    "All ext. cracks that are <10% CSD",
    "# cracks that are <50% CSD",
    "All ext. cracks that are <25% CSD",
    "Are there 2 or fewer cracks between 50-80% CSD",
    "All ext. cracks are <50% CSD",
    "One or more int. cracks that are >80% CSD",
    "Three or more int. cracks that are >50% CSD",
    "Any splits present",
    "OVERALL RATING",
]

RATING_THRESHOLDS = {
    "Rating 1": [
        "≤100% CSD",
        "Any number",
        "All <10%",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "Pass",
    ],
    "Rating 2": [
        "≤200% CSD",
        "-",
        "-",
        "Any number",
        "All <25%",
        "-",
        "-",
        "-",
        "-",
        "-",
        "Pass",
    ],
    "Rating 3": [
        "≤300% CSD",
        "-",
        "-",
        "-",
        "-",
        "≤2 cracks",
        "All <50%",
        "-",
        "-",
        "-",
        "Pass",
    ],
    "Rating 4": [
        "> 300% CSD",
        "-",
        "-",
        "-",
        "-",
        "-",
        "Any >50%",
        "≥1 crack >80%",
        "≥3 cracks >50%",
        "-",
        "Fail",
    ],
    "Rating 5": [
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        "Yes",
        "Fail",
    ],
}

RATING_HEADER_LABELS = ["Metric", "Value"] + list(RATING_THRESHOLDS.keys())


class MainWindow(QMainWindow):
    def __init__(self, debug_layout: bool = False):
        window_size = DEFAULT_LAYOUT["window"]["size"]
        super().__init__()
        self.setWindowTitle("oRinGD - ISO23936-2 Annex B Analyzer")
        self.resize(window_size[0], window_size[1])
        self.setMinimumSize(1000, 720)

        self.current_image_path: Optional[str] = None
        self._has_shown_crack_prompt = False
        self.session_records: List[SessionAnalysis] = []
        self.debug_layout = debug_layout
        self.settings_path = os.path.join(os.path.dirname(__file__), "layout_prefs.json")

        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)

        self.scene = CanvasScene()
        self.view = CanvasView(self.scene)
        self.view.setMinimumSize(700, 500)

        self.crack_table_widget = QTableWidget()
        self.crack_table_widget.setColumnCount(3)
        self.crack_table_widget.setHorizontalHeaderLabels(["Crack #", "Type", "Length, % of CSD"])
        crack_vheader = self.crack_table_widget.verticalHeader()
        if crack_vheader:
            crack_vheader.setVisible(False)
        self.crack_table_widget.setColumnWidth(0, 50)
        self.crack_table_widget.setColumnWidth(1, 75)
        crack_header = self.crack_table_widget.horizontalHeader()
        if crack_header:
            crack_header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.crack_table_widget.setMinimumWidth(220)

        self.top_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.top_splitter.addWidget(self.view)
        self.top_splitter.addWidget(self.crack_table_widget)
        self.top_splitter.setStretchFactor(0, 5)
        self.top_splitter.setStretchFactor(1, 2)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)

        rating_tab = QWidget()
        rating_layout = QVBoxLayout(rating_tab)
        rating_layout.setContentsMargins(0, 0, 0, 0)

        self.rating_table_widget = QTableWidget()
        self.rating_table_widget.setColumnCount(7)
        self.rating_table_widget.setHorizontalHeaderLabels(
            ["Metric", "Value", "Rating 1", "Rating 2", "Rating 3", "Rating 4", "Rating 5"]
        )
        rating_vheader = self.rating_table_widget.verticalHeader()
        if rating_vheader:
            rating_vheader.setVisible(False)
        self.rating_table_widget.setMinimumHeight(320)
        self.rating_table_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        if rating_vheader:
            rating_vheader.setDefaultSectionSize(26)
            rating_vheader.setMinimumSectionSize(24)
        rating_layout.addWidget(self.rating_table_widget)
        tabs.addTab(rating_tab, "Current Analysis")

        session_tab = QWidget()
        session_layout = QVBoxLayout(session_tab)
        session_layout.setContentsMargins(0, 0, 0, 0)

        session_label = QLabel("Session Summary")
        session_label.setStyleSheet("font-weight: bold;")
        session_layout.addWidget(session_label)

        self.session_table_widget = QTableWidget()
        self.session_table_widget.setMinimumHeight(160)
        self.session_table_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        session_vheader = self.session_table_widget.verticalHeader()
        if session_vheader:
            session_vheader.setDefaultSectionSize(24)
            session_vheader.setMinimumSectionSize(20)
        session_layout.addWidget(self.session_table_widget, stretch=1)

        session_actions_layout = QHBoxLayout()
        self.edit_session_button = QPushButton("Edit Selected")
        self.edit_session_button.clicked.connect(self.edit_selected_analysis)
        session_actions_layout.addWidget(self.edit_session_button)

        self.delete_session_button = QPushButton("Delete Selected")
        self.delete_session_button.clicked.connect(self.delete_selected_analysis)
        session_actions_layout.addWidget(self.delete_session_button)
        session_actions_layout.addStretch(1)
        session_layout.addLayout(session_actions_layout)

        tabs.addTab(session_tab, "Session Summary")
        self.tab_widget = tabs

        self.main_splitter = QSplitter(Qt.Orientation.Vertical)
        self.main_splitter.addWidget(self.top_splitter)
        self.main_splitter.addWidget(tabs)
        self.main_splitter.setStretchFactor(0, 5)
        self.main_splitter.setStretchFactor(1, 2)
        root_layout.addWidget(self.main_splitter, stretch=1)

        self.apply_layout_defaults()

        button_layout = QHBoxLayout()

        self.image_button = QPushButton("Load Image")
        self.image_button.clicked.connect(self.select_image)
        button_layout.addWidget(self.image_button)

        perim_mode_button = QPushButton("Perimeter Mode")
        perim_mode_button.clicked.connect(lambda: self.view.set_mode('draw_perimeter'))
        button_layout.addWidget(perim_mode_button)

        crack_mode_button = QPushButton("Crack Mode")
        crack_mode_button.clicked.connect(lambda: self.view.set_mode('draw_crack'))
        button_layout.addWidget(crack_mode_button)

        clear_button = QPushButton("Clear Active Analysis")
        clear_button.clicked.connect(self.clear_active_analysis)
        button_layout.addWidget(clear_button)

        self.save_report_button = QPushButton("Save Report")
        self.save_report_button.clicked.connect(self.saveAsExcel)
        button_layout.addWidget(self.save_report_button)

        debug_button = QPushButton("Debug Current Rating")
        debug_button.clicked.connect(self.debug_current_rating)
        button_layout.addWidget(debug_button)

        close_button = QPushButton("Close")
        close_button.clicked.connect(self.close)
        button_layout.addWidget(close_button)

        root_layout.addLayout(button_layout)

        self.initialize_rating_table()
        self.initialize_session_table()
        self.refresh_tables()

        self.view.perimeterUpdated.connect(self.refresh_tables)
        self.view.cracksUpdated.connect(self.refresh_tables)
        self.view.perimeterUpdated.connect(self.update_action_states)
        self.view.cracksUpdated.connect(self.update_action_states)
        self.view.modeChanged.connect(self.on_mode_changed)
        self.view.analysisFinalizeRequested.connect(self.finalize_current_analysis)

        if self.debug_layout:
            self.restore_layout_preferences()

        selection_model = self.session_table_widget.selectionModel()
        if selection_model:
            selection_model.selectionChanged.connect(lambda *_: self.update_action_states())

        self.show()
        self.update_action_states()
        self.show_perimeter_prompt()

    def on_mode_changed(self, mode: str):
        if mode == 'draw_perimeter':
            self._has_shown_crack_prompt = False
        if mode == 'draw_crack' and not self._has_shown_crack_prompt:
            self.show_crack_prompt()
            self._has_shown_crack_prompt = True

    def refresh_tables(self):
        self.update_crack_table()
        self.update_rating_table()
        self.update_action_states()

    def initialize_session_table(self):
        headers = ["#", "Image", "Completed", "Cracks", "Total % CSD", "Rating", "Result"]
        self.session_table_widget.setColumnCount(len(headers))
        self.session_table_widget.setHorizontalHeaderLabels(headers)
        sess_vheader = self.session_table_widget.verticalHeader()
        if sess_vheader:
            sess_vheader.setVisible(False)
        header = self.session_table_widget.horizontalHeader()
        if header:
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.session_table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.session_table_widget.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.session_table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.session_table_widget.setAlternatingRowColors(True)
        self.refresh_session_table()

    def refresh_session_table(self):
        self.session_table_widget.setRowCount(len(self.session_records))
        for row, record in enumerate(self.session_records):
            values = [
                str(record.index),
                record.image_name,
                record.completed_at.strftime("%Y-%m-%d %H:%M:%S"),
                str(record.crack_count),
                f"{record.total_pct:.2f}%",
                str(record.rating),
                record.result,
            ]
            for col, text in enumerate(values):
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 1:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.session_table_widget.setItem(row, col, item)

            result_item = self.session_table_widget.item(row, 6)
            if result_item:
                if record.result == "Pass":
                    result_item.setBackground(Qt.GlobalColor.green)
                    result_item.setForeground(Qt.GlobalColor.black)
                else:
                    result_item.setBackground(Qt.GlobalColor.red)
                    result_item.setForeground(Qt.GlobalColor.white)

        self.session_table_widget.resizeRowsToContents()

    def get_selected_session_row(self) -> Optional[int]:
        selection_model = self.session_table_widget.selectionModel()
        if not selection_model:
            return None
        rows = selection_model.selectedRows()
        if not rows:
            return None
        return rows[0].row()

    def delete_selected_analysis(self):
        row = self.get_selected_session_row()
        if row is None:
            QMessageBox.information(self, "Select Analysis", "Choose a session entry to delete.")
            return
        record = self.session_records[row]
        response = QMessageBox.question(
            self,
            "Delete Analysis?",
            f"Remove the recorded analysis for {record.image_name}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if response != QMessageBox.StandardButton.Yes:
            return
        self.session_records.pop(row)
        self.reindex_session_records()
        self.refresh_session_table()
        self.session_table_widget.clearSelection()
        self.update_action_states()

    def edit_selected_analysis(self):
        row = self.get_selected_session_row()
        if row is None:
            QMessageBox.information(self, "Select Analysis", "Choose a session entry to edit.")
            return
        record = self.session_records[row]

        if self.has_active_analysis_data():
            response = QMessageBox.question(
                self,
                "Replace Current Analysis?",
                "Current, unfinalized work will be discarded. Continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if response != QMessageBox.StandardButton.Yes:
                return

        if not os.path.exists(record.image_path):
            QMessageBox.warning(
                self,
                "Image Missing",
                f"Cannot find {record.image_path}. The session entry will be removed.",
            )
            self.session_records.pop(row)
            self.reindex_session_records()
            self.refresh_session_table()
            self.update_action_states()
            return

        if not self.view.load_image(record.image_path):
            QMessageBox.warning(self, "Load Failed", "Unable to load the selected image for editing.")
            return

        self.session_records.pop(row)
        self.reindex_session_records()
        self.refresh_session_table()
        self.session_table_widget.clearSelection()

        self.current_image_path = record.image_path
        self.view.set_mode('draw_perimeter')
        self._has_shown_crack_prompt = False
        self.refresh_tables()
        self.show_perimeter_prompt()
        QMessageBox.information(
            self,
            "Edit Mode",
            f"Recreate the perimeter and cracks for {record.image_name}, then finalize again when ready.",
        )
        self.update_action_states()

    def reindex_session_records(self):
        for idx, record in enumerate(self.session_records, start=1):
            record.index = idx

    def layout_preferences_payload(self) -> dict:
        return {
            "window": {
                "size": [self.width(), self.height()],
            },
            "top_splitter": self.top_splitter.sizes() if hasattr(self, "top_splitter") else [],
            "main_splitter": self.main_splitter.sizes() if hasattr(self, "main_splitter") else [],
        }

    def save_layout_preferences(self) -> None:
        if not self.debug_layout:
            return
        try:
            with open(self.settings_path, "w", encoding="utf-8") as fp:
                json.dump(self.layout_preferences_payload(), fp, indent=2)
        except OSError:
            pass

    def restore_layout_preferences(self) -> None:
        if not self.debug_layout:
            return
        if not os.path.exists(self.settings_path):
            return
        try:
            with open(self.settings_path, "r", encoding="utf-8") as fp:
                data = json.load(fp)
        except (OSError, json.JSONDecodeError):
            return

        window = data.get("window", {})
        size = window.get("size")
        if isinstance(size, list) and len(size) == 2:
            try:
                self.resize(int(size[0]), int(size[1]))
            except (TypeError, ValueError):
                pass

        top_sizes = data.get("top_splitter")
        if isinstance(top_sizes, list) and hasattr(self, "top_splitter"):
            self.top_splitter.setSizes([int(s) for s in top_sizes if isinstance(s, (int, float))])

        main_sizes = data.get("main_splitter")
        if isinstance(main_sizes, list) and hasattr(self, "main_splitter"):
            self.main_splitter.setSizes([int(s) for s in main_sizes if isinstance(s, (int, float))])

    def closeEvent(self, event):
        self.save_layout_preferences()
        super().closeEvent(event)

    def apply_layout_defaults(self):
        size = DEFAULT_LAYOUT.get("window", {}).get("size")
        if isinstance(size, list) and len(size) == 2:
            self.resize(int(size[0]), int(size[1]))
        top_sizes = DEFAULT_LAYOUT.get("top_splitter")
        if isinstance(top_sizes, list) and hasattr(self, "top_splitter"):
            self.top_splitter.setSizes(top_sizes)
        main_sizes = DEFAULT_LAYOUT.get("main_splitter")
        if isinstance(main_sizes, list) and hasattr(self, "main_splitter"):
            self.main_splitter.setSizes(main_sizes)

    def has_active_analysis_data(self) -> bool:
        perimeter = self.view.get_perimeter_data()
        _, cracks = self.view.engine_inputs()
        if len(perimeter.spline_points) >= 3 or len(perimeter.control_points) >= 3:
            return True
        return len(cracks) > 0

    def can_finalize_analysis(self) -> bool:
        if not self.current_image_path:
            return False
        perimeter = self.view.get_perimeter_data()
        return len(perimeter.spline_points) >= 3

    def update_action_states(self):
        if hasattr(self, "save_report_button"):
            self.save_report_button.setEnabled(bool(self.session_records))
        selection_model = self.session_table_widget.selectionModel() if hasattr(self, "session_table_widget") else None
        has_selection = bool(selection_model and selection_model.hasSelection())
        if hasattr(self, "edit_session_button"):
            self.edit_session_button.setEnabled(has_selection)
        if hasattr(self, "delete_session_button"):
            self.delete_session_button.setEnabled(has_selection)

    def update_crack_table(self):
        _, cracks = self.view.engine_inputs()
        self.crack_table_widget.setRowCount(len(cracks))
        for row, (crack_type, percent_length) in enumerate(cracks):
            number_item = QTableWidgetItem(str(row + 1))
            number_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            type_item = QTableWidgetItem(crack_type)
            type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            length_item = QTableWidgetItem(f"{percent_length:.2f}%")
            length_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            self.crack_table_widget.setItem(row, 0, number_item)
            self.crack_table_widget.setItem(row, 1, type_item)
            self.crack_table_widget.setItem(row, 2, length_item)

    def initialize_rating_table(self):
        self.rating_table_widget.setRowCount(len(RATING_METRICS))
        self.rating_table_widget.setColumnCount(len(RATING_HEADER_LABELS))
        self.rating_table_widget.setHorizontalHeaderLabels(RATING_HEADER_LABELS)

        vertical_header = self.rating_table_widget.verticalHeader()
        if vertical_header:
            vertical_header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.rating_table_widget.resizeColumnsToContents()
        header = self.rating_table_widget.horizontalHeader()
        if header:
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.rating_table_widget.setColumnWidth(0, 275)

        for row, metric in enumerate(RATING_METRICS):
            self.rating_table_widget.setItem(row, 0, QTableWidgetItem(metric))
            for col, threshold_key in enumerate(RATING_THRESHOLDS.keys(), start=2):
                self.rating_table_widget.setColumnWidth(col, 90)
                threshold_item = QTableWidgetItem(RATING_THRESHOLDS[threshold_key][row])
                threshold_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.rating_table_widget.setItem(row, col, threshold_item)

    def update_rating_table(self) -> None:
        if self.rating_table_widget.rowCount() == 0:
            self.initialize_rating_table()

        _, cracks = self.view.engine_inputs()
        metrics = compute_metrics(cracks)
        assigned_rating = assign_iso23936_rating(cracks)
        values = table_values(metrics)

        for row, text in enumerate(values):
            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.rating_table_widget.setItem(row, 1, item)

        for col in range(2, 7):
            for row in range(10):
                cell = self.rating_table_widget.item(row, col)
                if not cell:
                    continue
                if col == assigned_rating + 1:
                    cell.setBackground(Qt.GlobalColor.yellow)
                    cell.setForeground(Qt.GlobalColor.black)
                else:
                    cell.setData(Qt.ItemDataRole.BackgroundRole, None)
                    cell.setData(Qt.ItemDataRole.ForegroundRole, None)

        overall_row = 10
        overall_eval = "Pass" if assigned_rating <= 3 else "Fail"
        overall_text = f"Rating: {assigned_rating} - {overall_eval}"

        overall_item = QTableWidgetItem(overall_text)
        overall_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.rating_table_widget.setItem(overall_row, 1, overall_item)

        if overall_eval == "Pass":
            overall_item.setBackground(Qt.GlobalColor.green)
            overall_item.setForeground(Qt.GlobalColor.black)
        else:
            overall_item.setBackground(Qt.GlobalColor.red)
            overall_item.setForeground(Qt.GlobalColor.white)

        viewport = self.rating_table_widget.viewport()
        if viewport:
            viewport.update()

    def select_image(self):
        if self.current_image_path and self.has_active_analysis_data():
            response = QMessageBox.question(
                self,
                "Replace Current Analysis?",
                "Loading another image will discard the in-progress analysis. Continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if response != QMessageBox.StandardButton.Yes:
                return

        fname, _ = QFileDialog.getOpenFileName(
            self,
            "Load Image",
            "",
            "Images (*.png *.jpg *.jpeg *.bmp)",
        )
        if not fname:
            return
        if not self.view.load_image(fname):
            QMessageBox.warning(self, "Load Failed", "Failed to load the selected image.")
            return
        self.current_image_path = fname
        self.view.set_mode('draw_perimeter')
        self._has_shown_crack_prompt = False
        self.show_perimeter_prompt()
        self.update_action_states()

    def finalize_current_analysis(self):
        if not self.current_image_path:
            QMessageBox.warning(self, "No Image", "Load an image before finalizing an analysis.")
            return

        perimeter = self.view.get_perimeter_data()
        if len(perimeter.spline_points) < 3:
            QMessageBox.warning(self, "Incomplete Perimeter", "Define and confirm the perimeter before finalizing.")
            return

        _, cracks = self.view.engine_inputs()
        metrics = compute_metrics(cracks)
        rating = assign_iso23936_rating(cracks)
        result = "Pass" if rating <= 3 else "Fail"
        next_action = self._prompt_post_finalize_action(rating, result)
        if next_action == "continue":
            return

        self._store_finalized_analysis(cracks, metrics, rating, result)

        if next_action == "load":
            self.select_image()
        elif next_action == "report":
            self.saveAsExcel()

    def clear_active_analysis(self):
        response = QMessageBox.question(
            self,
            "Clear Active Analysis?",
            "Remove current perimeter and crack traces? Finalized analyses remain in the session table.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if response == QMessageBox.StandardButton.Yes:
            self.view.clear_overlays()
            self.view.set_mode('draw_perimeter')
            self._has_shown_crack_prompt = False
            self.refresh_tables()
            self.update_action_states()

    def saveCanvas(self, file_path: Optional[str] = None, suppress_conf: bool = False):
        if not file_path:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Image",
                "",
                "PNG Files (*.png);;All Files (*)",
            )
        if file_path:
            pixmap = self.view.grab()
            pixmap.save(file_path)
            if not suppress_conf:
                QMessageBox.information(self, "Success", f"Image saved to {file_path}")

    def _capture_view_snapshot(self) -> Optional[bytes]:
        try:
            pixmap = self.view.grab()
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                temp_path = tmp.name
            if not pixmap.save(temp_path, "PNG"):
                os.remove(temp_path)
                return None
            with open(temp_path, "rb") as fp:
                data = fp.read()
            os.remove(temp_path)
            return data
        except Exception:
            return None

    def _prompt_post_finalize_action(self, rating: int, result: str) -> Literal["load", "report", "continue"]:
        dialog = QMessageBox(self)
        dialog.setWindowTitle("Finalize Analysis")
        dialog.setText(f"Analysis will be recorded with Rating {rating} ({result}).")
        dialog.setInformativeText("What would you like to do next?")
        load_button = dialog.addButton("Load Another Image", QMessageBox.ButtonRole.AcceptRole)
        report_button = dialog.addButton("Produce Report", QMessageBox.ButtonRole.ActionRole)
        continue_button = dialog.addButton("Continue Drawing", QMessageBox.ButtonRole.RejectRole)
        dialog.setDefaultButton(load_button)
        dialog.exec()

        clicked = dialog.clickedButton()
        if clicked == load_button:
            return "load"
        if clicked == report_button:
            return "report"
        if clicked == continue_button:
            return "continue"
        return "continue"

    def _store_finalized_analysis(
        self,
        cracks: List[Crack],
        metrics,
        rating: int,
        result: str,
    ) -> SessionAnalysis:
        image_path = self.current_image_path or ""
        image_name = os.path.basename(image_path) if image_path else "Unknown Image"
        snapshot_png = self._capture_view_snapshot()

        record = SessionAnalysis(
            index=len(self.session_records) + 1,
            image_name=image_name,
            image_path=image_path,
            completed_at=datetime.datetime.now(),
            crack_count=metrics.num_cracks,
            total_pct=metrics.total_pct,
            rating=rating,
            result=result,
            cracks=list(cracks),
            snapshot_png=snapshot_png,
        )

        self.session_records.append(record)
        self.refresh_session_table()
        self.session_table_widget.scrollToBottom()

        self.view.clear_overlays()
        self.current_image_path = None
        self.refresh_tables()
        self.update_action_states()

        return record

    def saveAsExcel(self):
        if not self.session_records:
            QMessageBox.warning(self, "No Analyses", "Finalize at least one analysis before saving a report.")
            return

        current_date = datetime.datetime.now().strftime("%m%d%Y")
        default_report_name = f"session-report-{current_date}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Report",
            default_report_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )

        if not file_path:
            return

        workbook = Workbook()
        summary_sheet = cast(Optional[Worksheet], workbook.active)
        if summary_sheet is None:
            summary_sheet = workbook.create_sheet("Session Summary")
        else:
            summary_sheet.title = "Session Summary"
        self._populate_session_summary_sheet(summary_sheet)

        temp_image_paths: List[str] = []

        try:
            for record in self.session_records:
                snapshot_path = None
                if record.snapshot_png:
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        tmp.write(record.snapshot_png)
                        snapshot_path = tmp.name
                    temp_image_paths.append(snapshot_path)
                self._add_analysis_sheet(workbook, record, snapshot_path)

            workbook.save(file_path)
            QMessageBox.information(self, "Success", f"Report saved to {file_path}")
        finally:
            for tmp_path in temp_image_paths:
                if tmp_path and os.path.exists(tmp_path):
                    os.remove(tmp_path)

    def _populate_session_summary_sheet(self, sheet):
        sheet["A1"] = "Completed Analyses"
        headers = ["#", "Image", "Completed", "Cracks", "Total % CSD", "Rating", "Result"]
        for idx, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=idx, value=header)

        for offset, record in enumerate(self.session_records, start=1):
            row_idx = 2 + offset
            sheet.cell(row=row_idx, column=1, value=record.index)
            sheet.cell(row=row_idx, column=2, value=record.image_name)
            sheet.cell(row=row_idx, column=3, value=record.completed_at.strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(row=row_idx, column=4, value=record.crack_count)
            sheet.cell(row=row_idx, column=5, value=f"{record.total_pct:.2f}%")
            sheet.cell(row=row_idx, column=6, value=record.rating)
            sheet.cell(row=row_idx, column=7, value=record.result)

        analytics_widths = [6, 28, 20, 10, 14, 10, 10]
        for idx, width in enumerate(analytics_widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _add_analysis_sheet(self, workbook: Workbook, record: SessionAnalysis, snapshot_path: Optional[str]):
        base_title = f"{record.index:02d} - {os.path.splitext(record.image_name)[0]}"
        sheet_title = self._make_unique_sheet_title(workbook, base_title)
        sheet = workbook.create_sheet(sheet_title)

        if snapshot_path and os.path.exists(snapshot_path):
            excel_image = Image(snapshot_path)
            sheet.add_image(excel_image, "B2")

        metadata = [
            ("Image", record.image_name),
            ("Completed", record.completed_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Cracks", record.crack_count),
            ("Total % CSD", f"{record.total_pct:.2f}%"),
            ("Rating", record.rating),
            ("Result", record.result),
        ]

        for offset, (label, value) in enumerate(metadata, start=2):
            sheet.cell(row=offset, column=2, value=label)
            sheet.cell(row=offset, column=3, value=value)

        self._write_rating_table_to_sheet(sheet, record)
        self._write_crack_table_to_sheet(sheet, record)

    def _make_unique_sheet_title(self, workbook: Workbook, base_title: str) -> str:
        invalid_chars = set('[]:*?/\\')
        sanitized = ''.join('_' if c in invalid_chars else c for c in base_title).strip()
        sanitized = sanitized or "Analysis"
        sanitized = sanitized[:31]

        if sanitized not in workbook.sheetnames:
            return sanitized

        suffix = 2
        while True:
            suffix_text = f" ({suffix})"
            trimmed = sanitized[: 31 - len(suffix_text)]
            candidate = f"{trimmed}{suffix_text}"
            if candidate not in workbook.sheetnames:
                return candidate
            suffix += 1

    def _write_rating_table_to_sheet(self, sheet, record: SessionAnalysis, start_row: int = 2, start_col: int = 12):
        header_row = start_row
        for offset, header in enumerate(RATING_HEADER_LABELS):
            sheet.cell(row=header_row, column=start_col + offset, value=header)

        metrics = compute_metrics(record.cracks)
        values = table_values(metrics)
        overall_text = f"Rating: {record.rating} - {'Pass' if record.result == 'Pass' else 'Fail'}"

        for metric_idx, metric_label in enumerate(RATING_METRICS):
            row_idx = header_row + 1 + metric_idx
            sheet.cell(row=row_idx, column=start_col, value=metric_label)

            if metric_idx < len(values):
                sheet.cell(row=row_idx, column=start_col + 1, value=values[metric_idx])
            else:
                sheet.cell(row=row_idx, column=start_col + 1, value=overall_text)

            for col_offset, threshold_key in enumerate(RATING_THRESHOLDS.keys(), start=2):
                sheet.cell(
                    row=row_idx,
                    column=start_col + col_offset,
                    value=RATING_THRESHOLDS[threshold_key][metric_idx],
                )

        for idx in range(len(RATING_HEADER_LABELS)):
            col_letter = get_column_letter(start_col + idx)
            sheet.column_dimensions[col_letter].width = 20 if idx else 50

    def _write_crack_table_to_sheet(self, sheet, record: SessionAnalysis, start_row: int = 16, start_col: int = 12):
        headers = ["Crack #", "Type", "Length (% of CSD)"]
        for offset, header in enumerate(headers):
            sheet.cell(row=start_row, column=start_col + offset, value=header)

        for row_offset, (crack_type, length_pct) in enumerate(record.cracks, start=1):
            sheet.cell(row=start_row + row_offset, column=start_col, value=row_offset)
            sheet.cell(row=start_row + row_offset, column=start_col + 1, value=crack_type)
            sheet.cell(row=start_row + row_offset, column=start_col + 2, value=f"{length_pct:.2f}%")

    def debug_current_rating(self):
        perimeter = self.view.get_perimeter_data()
        if len(perimeter.spline_points) < 3:
            QMessageBox.warning(self, "No Perimeter", "Please define a perimeter first.")
            return

        debug_info = self.get_rating_debug_info()

        dialog = QDialog(self)
        dialog.setWindowTitle("Rating Debug Information")
        dialog.setMinimumSize(500, 800)

        layout = QVBoxLayout()

        text_edit = QTextEdit()
        text_edit.setPlainText(debug_info)
        text_edit.setReadOnly(True)
        layout.addWidget(text_edit)

        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        layout.addWidget(close_button)

        dialog.setLayout(layout)
        dialog.exec()

    def get_rating_debug_info(self) -> str:
        per = self.view.get_perimeter_data()
        if len(per.spline_points) < 3:
            return "No perimeter defined"

        csd_px, cracks = self.view.engine_inputs()

        info = ["CURRENT RATING DEBUG INFORMATION", "=" * 40]
        info.append(f"CSD: {csd_px:.2f} pixels")
        info.append(f"Number of cracks: {len(cracks)}")
        info.append("")

        total_length_percent = sum(length for _, length in cracks)
        internal_lengths = [length for ctype, length in cracks if ctype == "Internal"]
        external_lengths = [length for ctype, length in cracks if ctype == "External"]
        has_split = any(ctype == "Split" for ctype, _ in cracks)

        for idx, (ctype, length) in enumerate(cracks, start=1):
            info.append(f"Crack {idx}: {ctype}, {length:.2f}% CSD")

        info.append("")
        info.append("DETAILED METRICS FOR RATING DETERMINATION:")
        info.append("-" * 40)

        all_crack_lengths = internal_lengths + external_lengths

        info.append("Rating 1 Requirements:")
        all_below_25 = all(length < 25 for length in all_crack_lengths) if all_crack_lengths else True
        all_ext_below_10 = all(length < 10 for length in external_lengths) if external_lengths else True
        num_at_or_above_25 = sum(1 for length in all_crack_lengths if length >= 25)
        info.append(f"  Total length ≤100%: {total_length_percent:.2f}% ({'✓' if total_length_percent <= 100 else '✗'})")
        info.append(f"  All cracks <25%: {'✓' if all_below_25 else '✗'}")
        if not all_below_25:
            info.append(f"    → {num_at_or_above_25} crack(s) ≥25% CSD")
        info.append(f"  All external <10%: {'✓' if all_ext_below_10 else '✗'}")

        info.append("\nRating 2 Requirements:")
        all_below_50 = all(length < 50 for length in all_crack_lengths) if all_crack_lengths else True
        all_ext_below_25 = all(length < 25 for length in external_lengths) if external_lengths else True
        num_at_or_above_50 = sum(1 for length in all_crack_lengths if length >= 50)
        info.append(f"  Total length ≤200%: {total_length_percent:.2f}% ({'✓' if total_length_percent <= 200 else '✗'})")
        info.append(f"  All cracks <50%: {'✓' if all_below_50 else '✗'}")
        if not all_below_50:
            info.append(f"    → {num_at_or_above_50} crack(s) ≥50% CSD")
        info.append(f"  All external <25%: {'✓' if all_ext_below_25 else '✗'}")

        info.append("\nRating 3 Requirements:")
        internal_50_80_count = sum(1 for length in internal_lengths if 50 <= length <= 80)
        all_ext_below_50 = all(length < 50 for length in external_lengths) if external_lengths else True
        info.append(f"  Total length ≤300%: {total_length_percent:.2f}% ({'✓' if total_length_percent <= 300 else '✗'})")
        info.append(f"  ≤2 internal cracks 50-80%: {internal_50_80_count} ({'✓' if internal_50_80_count <= 2 else '✗'})")
        info.append(f"  All external <50%: {'✓' if all_ext_below_50 else '✗'}")

        info.append("\nRating 4 Triggers (any one triggers failure):")
        internal_above_80_count = sum(1 for length in internal_lengths if length > 80)
        internal_above_50_count = sum(1 for length in internal_lengths if length > 50)
        any_ext_above_50 = any(length > 50 for length in external_lengths) if external_lengths else False
        info.append(f"  Total >300%: {'✗' if total_length_percent > 300 else '✓'}")
        info.append(f"  ≥1 internal >80%: {internal_above_80_count} ({'✗' if internal_above_80_count >= 1 else '✓'})")
        info.append(f"  ≥3 internals >50%: {internal_above_50_count} ({'✗' if internal_above_50_count >= 3 else '✓'})")
        info.append(f"  Any external >50%: {'✗' if any_ext_above_50 else '✓'}")

        info.append("\nRating 5 Trigger:")
        info.append(f"  Any split present: {'✗' if has_split else '✓'}")

        info.append("\n" + "=" * 40)
        info.append("RATING DECISION TREE:")

        if not cracks:
            assigned_rating = 0
            info.append("No cracks → Rating 0")
        elif has_split:
            assigned_rating = 5
            info.append("Has split → Rating 5 (FAIL)")
        else:
            if total_length_percent <= 100 and all_below_25 and all_ext_below_10:
                assigned_rating = 1
                info.append("Meets Rating 1 conditions")
            elif (
                total_length_percent > 300
                or internal_above_80_count >= 1
                or internal_above_50_count >= 3
                or any_ext_above_50
            ):
                assigned_rating = 4
                info.append("Triggers Rating 4 conditions (FAIL)")
                triggers = []
                if total_length_percent > 300:
                    triggers.append(f"Total {total_length_percent:.1f}% > 300%")
                if internal_above_80_count >= 1:
                    triggers.append(f"{internal_above_80_count} internal(s) > 80%")
                if internal_above_50_count >= 3:
                    triggers.append(f"{internal_above_50_count} internals > 50%")
                if any_ext_above_50:
                    triggers.append("External crack > 50%")
                info.append(f"   Triggered by: {', '.join(triggers)}")
            elif total_length_percent <= 200 and all_below_50 and all_ext_below_25:
                assigned_rating = 2
                info.append("Meets Rating 2 conditions")
            elif total_length_percent <= 300 and internal_50_80_count <= 2 and all_ext_below_50:
                assigned_rating = 3
                info.append("Meets Rating 3 conditions")
            else:
                assigned_rating = 4
                info.append("Default to Rating 4 (no other conditions met)")

        info.append(f"\nFINAL RATING: {assigned_rating} - {'PASS' if assigned_rating <= 3 else 'FAIL'}")
        return "\n".join(info)

    def show_perimeter_prompt(self):
        QMessageBox.information(
            self,
            "Set Perimeter",
            "Use the left-mouse button to add points around the perimeter of the o-ring.\n\n"
            "When finished, click the middle mouse button to generate the loop and review the line fit.\n\n"
            "You may use the right-mouse button to delete any points or the fitted perimeter line before confirming.\n\n"
            "Once happy, click the middle mouse button again to confirm.\n",
        )

    def show_crack_prompt(self):
        QMessageBox.information(
            self,
            "Trace cracks",
            "Click and drag the left-mouse button to trace the visible cracks on the o-ring.\n\n"
            "Release the left-mouse button to confirm a crack and add it to the analysis.\n\n"
            "You may use the right-mouse button to delete any number of drawn cracks from the analysis.\n\n"
            "You can click the 'Clear Session' button to re-start the analysis.\n",
        )


def main(argv: Optional[list[str]] = None):
    parser = argparse.ArgumentParser(description="oRinGD - ISO23936-2 Annex B Analyzer")
    parser.add_argument(
        "--debug-layout",
        action="store_true",
        help="Enable saving/loading layout_prefs.json for splitter/window sizing",
    )
    args = parser.parse_args(argv)

    app = QApplication(sys.argv)
    window = MainWindow(debug_layout=args.debug_layout)
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
